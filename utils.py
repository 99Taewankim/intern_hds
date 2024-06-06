def clear_dir(path):
    '''
        폴더 비우기
        Input:
            path - (str) 비울 폴더 경로
    '''
    import os
    import shutil
    from pdb import set_trace as st

    shutil.rmtree(path, ignore_errors=True)
    os.makedirs(path, exist_ok=True)


def read_info_from_excel(fileName="info.xlsx"):
    '''
    info.xlsx 파일 읽기
    '''
    from openpyxl import load_workbook
    
    wb = load_workbook(fileName, data_only=True)
    ws = wb.active

    info = {}
    for i in range(1, ws.max_row):
        row = []
        for j in range(0, ws.max_column):
            row.append(ws.cell(i+1, j+1).value)

        label = ws.cell(i+1,2).value
        key = ws.cell(i+1,4).value
        value = ws.cell(i+1,5).value

        if label not in info:
            info[label] = {}    
        info[label][key] = value

    return info

def to_valid_file_name(name):
    '''
        파일/폴더명으로 사용 불가능한 문자 삭제하기
    '''
    import re
    
    BANNED_LETTER = r'/|\\|:|[*]|[?]|<|>|\|'
    BANNED_NAME = ['con', 'aux', 'nul', 'prn',
                   'com0', 'com1', 'com2', 'com3', 'com4',
                   'com5', 'com6', 'com7', 'com8', 'com9',
                   'lpt0', 'lpt1', 'lpt2', 'lpt3', 'lpt4',
                   'lpt5', 'lpt6', 'lpt7', 'lpt8', 'lpt9']
    
    validName = name
    
    # 사용 불가능한 문자 지우기
    validName = re.sub(BANNED_LETTER, '', validName)
    
    # 폴더명으로 사용 불가능한 이름일 경우
    # -> 언더바 추가
    if validName in BANNED_NAME:
        validName = f'{validName}_'
    
    return validName